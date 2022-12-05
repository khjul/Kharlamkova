import csv
import math
import re
import matplotlib.pyplot as plt
import numpy as np
import wkhtmltopdf
from jinja2 import Environment, FileSystemLoader
import pdfkit

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055}

class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = list()

    def get_dataset(file_name):
        dataset = DataSet(file_name)
        dictionary_list = DataSet.csv_filter(DataSet.csv_reader(file_name)[0], DataSet.csv_reader(file_name)[1])
        for vacancy in dictionary_list:
            vacancy = Vacancy([f"{vacancy['name']}", f"{vacancy['salary_from']}",
                               f"{vacancy['salary_to']}",
                               f"{vacancy['salary_currency']}",
                               f"{vacancy['area_name']}",
                               f"{vacancy['published_at']}"])
            vacancy.published_at = int(vacancy.published_at[:4])
            dataset.vacancies_objects.append(vacancy)
        return dataset

    def csv_reader(file_name):
        with open(file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            data = [line for line in reader]
        if len(data) == 0:
            print("Пустой файл")
            exit(0)
        headings = data[0]
        line = data[1:]
        return headings, line

    def string_filter(str_with_tags):
        result = re.sub("<.*?>", '', str_with_tags)
        if ('\n' in str_with_tags):
            return result
        else:
            return " ".join(result.split())

    def csv_filter(headings, lines):
        def check(line):
            return (len(line) == len(headings) and line.count('') == 0)

        vacancies_list = list(filter(check, lines))
        dictionary_vacancies_list = [dict(zip(headings, map(DataSet.string_filter, line))) for line in vacancies_list]
        return dictionary_vacancies_list

class Vacancy:
    def __init__(self, args):
        self.name = args[0]
        self.salary_from = float(args[1])
        self.salary_to = float(args[2])
        self.salary_currency = args[3]
        self.area_name = args[4]
        self.published_at = args[5]

class InputConnect:
    def __init__(self, input_file_name, input_profession_name):
        self.file_name_init = input_file_name
        self.prof_name_init = input_profession_name

    def get_right_course(vacancy):
        salary = int((vacancy.salary_from * currency_to_rub[vacancy.salary_currency]
                      + vacancy.salary_to * currency_to_rub[vacancy.salary_currency]) / 2)
        return salary

    def get_count_vacancies(data: DataSet, parametr):
        count_vacancies = dict()
        for vacancy in data.vacancies_objects:
            if parametr == "all":
                if count_vacancies.__contains__(vacancy.published_at):
                    count_vacancies[vacancy.published_at] += 1
                else:
                    count_vacancies[vacancy.published_at] = 1
        if (len(count_vacancies) == 0):
            return {2022: 0}
        return count_vacancies

    def get_count_vacancies_by_profession(data: DataSet, profession):
        count_vacancies_by_profession = dict()
        for vacancy in data.vacancies_objects:
            if vacancy.name.__contains__(profession):
                if not count_vacancies_by_profession.__contains__(vacancy.published_at):
                    count_vacancies_by_profession[vacancy.published_at] = 1
                else:
                    count_vacancies_by_profession[vacancy.published_at] += 1
        if (len(count_vacancies_by_profession) == 0):
            return {2022: 0}
        return count_vacancies_by_profession

    def get_general_salary_level_by_year(data: DataSet, parametr):
        general_salary_level_by_year = dict()
        for vac in data.vacancies_objects:
            if general_salary_level_by_year.__contains__(vac.published_at):
                general_salary_level_by_year[vac.published_at] += InputConnect.get_right_course(vac)
            else:
                general_salary_level_by_year[vac.published_at] = InputConnect.get_right_course(vac)
        if (len(general_salary_level_by_year) == 0):
            return {2022: 0}
        for key in general_salary_level_by_year.keys():
            if parametr == "all":
                general_salary_level_by_year[key] = math.floor(
                    general_salary_level_by_year[key] / data.general_count_vacancies_by_year[key])
        return general_salary_level_by_year

    def get_salary_level_by_profession(data: DataSet, profession):
        salary_level_by_profession = dict()
        for vacancy in data.vacancies_objects:
            if vacancy.name.__contains__(profession):
                if salary_level_by_profession.__contains__(vacancy.published_at):
                    salary_level_by_profession[vacancy.published_at] += InputConnect.get_right_course(vacancy)
                else:
                    salary_level_by_profession[vacancy.published_at] = InputConnect.get_right_course(vacancy)
        if (len(salary_level_by_profession) == 0):
            return {2022: 0}
        for key in salary_level_by_profession.keys():
            if profession != "all":
                salary_level_by_profession[key] = math.floor(
                    salary_level_by_profession[key] / data.count_vacancies_by_profession[key])
        return salary_level_by_profession

    def get_proportion_vacancy_by_cities(data: DataSet):
        proportion_vacancy_by_cities = dict()
        for vacancy in data.vacancies_objects:
            if proportion_vacancy_by_cities.__contains__(vacancy.area_name):
                proportion_vacancy_by_cities[vacancy.area_name] += 1
            else:
                proportion_vacancy_by_cities[vacancy.area_name] = 1
        return proportion_vacancy_by_cities

    def get_salary_level_by_cities(data: DataSet):
        salary_level_by_cities = dict()
        for vacancy in data.vacancies_objects:
            if math.floor(
                    data.proportion_vacancy_by_cities[vacancy.area_name] / len(data.vacancies_objects) * 100) >= 1:
                if salary_level_by_cities.__contains__(vacancy.area_name):
                    salary_level_by_cities[vacancy.area_name] += InputConnect.get_right_course(vacancy)
                else:
                    salary_level_by_cities[vacancy.area_name] = InputConnect.get_right_course(vacancy)
        for key in salary_level_by_cities:
            salary_level_by_cities[key] = math.floor(
                salary_level_by_cities[key] / data.proportion_vacancy_by_cities[key])
        result = {k: v for k, v in sorted(salary_level_by_cities.items(), key=lambda item: item[1], reverse=True)}
        return result

    def get_right_proportion_vacancy_by_cities(data: DataSet):
        data.proportion_vacancy_by_cities = {k: round(v / len(data.vacancies_objects), 4)
                                             for k, v in data.proportion_vacancy_by_cities.items()}
        data.proportion_vacancy_by_cities = {k: v for k, v in data.proportion_vacancy_by_cities.items()
                                             if math.floor(v * 100 >= 1)}
        return {k: v for k, v in
                sorted(data.proportion_vacancy_by_cities.items(), key=lambda item: item[1], reverse=True)}

    def print(self, data: DataSet):
        data.general_count_vacancies_by_year = InputConnect.get_count_vacancies(data, "all")
        data.general_salary_level_by_year = InputConnect.get_general_salary_level_by_year(data, "all")

        data.count_vacancies_by_profession = InputConnect.get_count_vacancies_by_profession(data, self.prof_name_init)
        data.salary_level_by_profession = InputConnect.get_salary_level_by_profession(data, self.prof_name_init)

        data.proportion_vacancy_by_cities = InputConnect.get_proportion_vacancy_by_cities(data)
        data.salary_level_by_cities = InputConnect.get_salary_level_by_cities(data)

        data.salary_level_by_cities_first_ten = dict(list(data.salary_level_by_cities.items())[:10])

        data.proportion_vacancy_by_cities = InputConnect.get_right_proportion_vacancy_by_cities(data)
        data.proportion_vacancy_by_cities_first_ten = dict(list(data.proportion_vacancy_by_cities.items())[:10])

        return data.general_salary_level_by_year, data.general_count_vacancies_by_year, data.salary_level_by_profession, data.count_vacancies_by_profession, data.salary_level_by_cities_first_ten, data.proportion_vacancy_by_cities_first_ten

class Report(InputConnect):
    def __init__(self, dict1, dict2, dict3, dict4 ,dict5, dict6):
        self.general_salary_level_by_year = dict1
        self.general_count_vacancies_by_year = dict2
        self.salary_level_by_profession = dict3
        self.count_vacancies_by_profession = dict4
        self.salary_level_by_cities_first_ten = dict5
        self.proportion_vacancy_by_cities_first_ten = dict6

    def as_text(val):
        if val is None:
            return ""
        return str(val)

    def generate_excel(self):
        wb = Workbook()
        sheet1 = wb.active
        thin = Side(border_style='thin', color='000000')
        sheet1.title = 'Статистика по годам'
        sheet2 = wb.create_sheet('Статистика по городам')
        prof = input_profession_name

        border = Border(left=thin, top=thin, right=thin, bottom=thin)

        heads1 = ['Год', 'Динамика уровня зарплат по годам',
                  f'Динамика уровня зарплат по годам для выбранной профессии - {prof}',
                  'Динамика количества вакансий по годам', f'Динамика количества вакансий по годам для выбранной профессии - {prof}']

        for i, head in enumerate(heads1):
            sheet1.cell(row = 1, column = (i + 1), value = head).font = Font(bold=True)

        year = list(self.general_salary_level_by_year.keys())[0]
        for i in range(len(self.general_salary_level_by_year)):
            sheet1.append([year, self.general_salary_level_by_year[year], self.general_count_vacancies_by_year[year],
                           self.salary_level_by_profession[year], self.count_vacancies_by_profession[year]])
            year += 1

        for column in sheet1.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        for column in sheet1.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet1.column_dimensions[column[1].column_letter].width = length + 2

        heads2 = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий по городам']

        for i, head in enumerate(heads2):
            sheet2.cell(row = 1, column = (i + 1), value = head).font = Font(bold=True)

        for i in range(len(self.salary_level_by_cities_first_ten)):
            city = list(self.salary_level_by_cities_first_ten.keys())[i]
            sheet2.append([city, self.salary_level_by_cities_first_ten[city]])

        j = 2
        for city in self.proportion_vacancy_by_cities_first_ten.keys():
            sheet2[f'D{j}'].value = city
            sheet2[f'E{j}'].value = f'{round(self.proportion_vacancy_by_cities_first_ten[city] * 100, 2)}%'
            sheet2[f'E{j}'].alignment = openpyxl.styles.Alignment(horizontal='right')
            j += 1

        for column in sheet2.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 2

        for i in range(1, 12):
            sheet2[f'A{i}'].border = border
            sheet2[f'B{i}'].border = border
            sheet2[f'D{i}'].border = border
            sheet2[f'E{i}'].border = border

        wb.save('report1.xlsx')

    def generate_image(self):
        fig = plt.figure()

        ax_1 = fig.add_subplot(2, 2, 1)
        ax_2 = fig.add_subplot(2, 2, 2)
        ax_3 = fig.add_subplot(2, 2, 3)
        ax_4 = fig.add_subplot(2, 2, 4)

        labels = self.general_salary_level_by_year.keys()

        profession = input_profession_name
        x = np.arange(len(labels))

        ax_1.set_title('Уровень зарплат по годам', fontsize=8)
        ax_1.bar(x, self.general_salary_level_by_year.values(), width = 0.4, label="средняя з/п")
        ax_1.bar(x + 0.4, self.salary_level_by_profession.values(), width = 0.4, label=f'з/п {profession.lower()}')
        ax_1.set_xticks(x, labels, rotation = 90)
        ax_1.legend(fontsize=8, loc=2)
        ax_1.tick_params(axis='both', labelsize=8)
        ax_1.grid(axis='y')

        ax_2.set_title('Количество вакансий по годам', fontsize=8)
        ax_2.bar(x, self.general_count_vacancies_by_year.values(), width = 0.4, label = "Количество вакансий")
        ax_2.bar(x + 0.4, self.count_vacancies_by_profession.values(), width = 0.4, label = f'Количество вакансий {profession.lower()}')
        ax_2.set_xticks(x, labels, rotation = 90)
        ax_2.legend(fontsize = 8, loc = 2)
        ax_2.tick_params(axis='both', labelsize = 8)
        ax_2.grid(axis='y')

        cities = [city.replace('-', '-\n').replace(' ', '\n') for city in self.salary_level_by_cities_first_ten.keys()]

        ax_3.set_title('Уровень зарплат по городам', fontsize=8)
        ax_3.barh(cities, list(reversed(self.salary_level_by_cities_first_ten.values())), height=0.8, align = 'center')
        ax_3.xaxis.set_tick_params(labelsize = 8)
        ax_3.yaxis.set_tick_params(labelsize = 6)
        ax_3.grid(axis='x')

        labels_cities =  list(self.proportion_vacancy_by_cities_first_ten.keys()) + ['Другие']
        others_sizes = [1 - sum(self.proportion_vacancy_by_cities_first_ten.values())]

        ax_4.set_title('Доля вакансий по городам', fontsize=8)
        ax_4.pie(list(self.proportion_vacancy_by_cities_first_ten.values()) + others_sizes, labels=labels_cities, textprops={'fontsize': 6})
        ax_4.axis('equal')

        plt.tight_layout()
        plt.savefig('graph1.png')

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('../homework python'))
        template = env.get_template('pdf.html')

        profession = input_profession_name
        image_file = 'C:\\Users\\kh_ju\\homework python\\graph.png'

        rows = []
        for year in general_salary_level_by_year.keys():
            rows.append([year, general_salary_level_by_year[year], general_count_vacancies_by_year[year],
                         salary_level_by_profession[year], count_vacancies_by_profession[year]])

        for key, value in proportion_vacancy_by_cities_first_ten.items():
            value = f'{round(value * 100, 2)}%'
            proportion_vacancy_by_cities_first_ten[key] = value

        pdf_template = template.render({'profession': profession, 'image_file': image_file, 'rows': rows,
                                        'general_salary_level_by_year': general_salary_level_by_year,
                                        'general_count_vacancies_by_year': general_count_vacancies_by_year,
                                        'salary_level_by_profession': salary_level_by_profession,
                                        'count_vacancies_by_profession': count_vacancies_by_profession,
                                        'salary_level_by_cities_first_ten': salary_level_by_cities_first_ten,
                                        'proportion_vacancy_by_cities_first_ten': proportion_vacancy_by_cities_first_ten})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Users\kh_ju\homework python\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

input_file_name = input("Введите название файла: ")
input_profession_name = input("Введите название профессии: ")
user_input = input("Вакансии или статистика? ").lower()

dataset_vacancies = DataSet.get_dataset(file_name=input_file_name)
InputConnect.print(InputConnect(input_file_name=input_file_name, input_profession_name=input_profession_name),
                   dataset_vacancies)

dicts = InputConnect.print(InputConnect(input_file_name=input_file_name, input_profession_name=input_profession_name),
                           dataset_vacancies)

general_salary_level_by_year = dicts[0]
general_count_vacancies_by_year = dicts[1]
salary_level_by_profession = dicts[2]
count_vacancies_by_profession = dicts[3]
salary_level_by_cities_first_ten = dicts[4]
proportion_vacancy_by_cities_first_ten = dicts[5]

report = Report(dict1=general_salary_level_by_year, dict2=general_count_vacancies_by_year,
                             dict3=salary_level_by_profession, dict4=count_vacancies_by_profession,
                             dict5=salary_level_by_cities_first_ten, dict6=proportion_vacancy_by_cities_first_ten)



if user_input == 'вакансии':
    Report.generate_image(report)
elif user_input == 'статистика':
    Report.generate_excel(report)




