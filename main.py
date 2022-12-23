import csv
import math
import re
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
import cProfile
from time import strptime
from _datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055}

class DataSet:
    """Класс для получения данных из CSV файла.

    Attriburies:
        file_name (str): название файла
        vacancies_objects (list): список со всеми вакансиями и информацией по ним
    """
    def __init__(self, file_name):
        """Инициализирует объект DataSet

        Args:
            file_name (str): название файла
            vacancies_objects (list): список со всеми вакансиями и информацией по ним
        """
        self.file_name = file_name
        self.vacancies_objects = list()

    def get_year(published_at):
        return int(published_at[:4])

    # def get_year_1(published_at):
    #     return datetime.strptime(published_at, "%Y-%m-%dT%H:%M:%S%z").year

    # def get_year_2(published_at):
    #     return datetime(int(published_at[:4]), int(published_at[5:7]), int(published_at[8:10])).year

    def get_dataset(file_name):
        """Считывает и фильтрует CSV файл, формирует объекты класса Vacancy, добавляет эти объекты в список

        :param file_name: название файла (str)
        :return:
            DataSet: объект класса DataSet

        >>> dataset = DataSet.get_dataset("vacancies_by_year.csv")
        >>> type(dataset.vacancies_objects[0]).__name__
        'Vacancy'

        >>> dataset = DataSet.get_dataset("vacancies_by_year.csv")
        >>> len(dataset.vacancies_objects)
        927145

        >>> dataset = DataSet.get_dataset("vacancies_by_year.csv")
        >>> dataset.vacancies_objects[0].salary_from
        35000.0

        >>> dataset = DataSet.get_dataset("vacancies_by_year.csv")
        >>> dataset.vacancies_objects[0].salary_to
        45000.0

        >>> dataset = DataSet.get_dataset("vacancies_by_year.csv")
        >>> dataset.vacancies_objects[0].published_at
        2007
        """
        dataset = DataSet(file_name)
        dictionary_list = DataSet.csv_filter(DataSet.csv_reader(file_name)[0], DataSet.csv_reader(file_name)[1])
        for vacancy in dictionary_list:
            vacancy = Vacancy([f"{vacancy['name']}", f"{vacancy['salary_from']}",
                               f"{vacancy['salary_to']}",
                               f"{vacancy['salary_currency']}",
                               f"{vacancy['area_name']}",
                               f"{vacancy['published_at']}"])
            vacancy.published_at = DataSet.get_year(vacancy.published_at)
            dataset.vacancies_objects.append(vacancy)
        return dataset

    def csv_reader(file_name):
        """Считывает CSV файл и создает список с заголовками и список с объектами Vacancy

        :return:
            str: Список с заголовками
            list: Список списков с информацией о вакансиях
        >>> data = DataSet.csv_reader("vacancies_by_year.csv")
        >>> len(data[0])
        6

        >>> data = DataSet.csv_reader("vacancies_by_year.csv")
        >>> len(data[1])
        927145
        """
        with open(file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            data = [line for line in reader]
        if len(data) == 0:
            print("Пустой файл")
            exit(0)
        headings = data[0]
        lines = data[1:]
        return headings, lines

    def string_filter(str_with_tags):
        """Очищает строку от HTML-тэгов.

        :param str_with_tags: строка (str)
        :return:
            str: строка без HTML-тэгов

        >>> DataSet.string_filter('<h1> string </h1>')
        'string'
        >>> DataSet.string_filter('<br> second string')
        'second string'
        >>> DataSet.string_filter('third <br> string')
        'third string'
        """

        result = re.sub("<.*?>", '', str_with_tags)
        if ('\n' in str_with_tags):
            return result
        else:
            return " ".join(result.split())

    def csv_filter(headings, lines):
        """Фильтрует и формирует список из строк файла

        :param headings: Список заголовков CSV файла
        :param lines: Список строк с информацией о вакансиях
        :return:
            list: Список словараей, где
                                        ключ: заголовок
                                        значение: значение в строке под этим заголовком
        """
        def check(line):
            """Проверяет на совпадение количества элементов строки с количеством заголовоков
                и на наличие пустых элементов в строке.
            :param line: строка из файла
            :return:
                bool: True or False
            """
            return (len(line) == len(headings) and line.count('') == 0)

        vacancies_list = list(filter(check, lines))
        dictionary_vacancies_list = [dict(zip(headings, map(DataSet.string_filter, line))) for line in vacancies_list]
        return dictionary_vacancies_list

class Vacancy:
    """Класс для представления вакансии.

    Attributes:
        name (str): Название вакансии
        salary_from (float): Нижняя граница оклада
        salary_to (float): Верхняя граница оклада
        salary_currency (str): Валюта
        area_name (str): Название региона
        published_at (str): Дата публикации вакансии
    """
    def __init__(self, args):
        """Инициализирует объект Vacancy.

        :param args: список строк с информацией о вакансии
        """
        self.name = args[0]
        self.salary_from = float(args[1])
        self.salary_to = float(args[2])
        self.salary_currency = args[3]
        self.area_name = args[4]
        self.published_at = args[5]

class InputConnect:
    """Отвечает за обработку параметров вводимых пользователем и формирует отчет."""
    def __init__(self, input_file_name, input_profession_name):
        """Инициализирует объект InputConnect.

        :param input_file_name: введенное пользователем имя файла
        :param input_profession_name: введенное пользователем название профессии

        >>> type(InputConnect('vacancies_by_year.csv', 'Программист')).__name__
        'InputConnect'
        >>> InputConnect('vacancies_by_year.csv', 'Программист').file_name_init
        'vacancies_by_year.csv'
        >>> InputConnect('vacancies_by_year.csv', 'Программист').prof_name_init
        'Программист'
        """

        self.file_name_init = input_file_name
        self.prof_name_init = input_profession_name

    def get_right_course(vacancy):
        """С помощью словаря currency_to_rub переводит зарплату в рубли и вычисляет среднюю зарплату.

        :param vacancy: объект класса Vacancy
        :return
            int: средняя зарплата

        >>> InputConnect.get_right_course(Vacancy(args=['Программист баз данных', '36000', '50000', 'RUR', 'Москва', '2007']))
        43000
        >>> InputConnect.get_right_course(Vacancy(args=['Программист баз данных', '40000', '80000', 'EUR', 'Москва', '2007']))
        3594000
        >>> InputConnect.get_right_course(Vacancy(args=['Программист баз данных', '30000', '60000', 'UAH', 'Москва', '2007']))
        73800

        """
        salary = int((vacancy.salary_from * currency_to_rub[vacancy.salary_currency]
                      + vacancy.salary_to * currency_to_rub[vacancy.salary_currency]) / 2)
        return salary

    def get_count_vacancies(data: DataSet, parametr):
        """Считает количество всех вакансий.

        :param parametr: параметр, определяющий необходимость подсчета количества всех вакансий
        :return:
            dict: словарь, где
                                ключ: год
                                значение: количество вакансий за этот год
        """
        count_vacancies = dict()
        for vacancy in data.vacancies_objects:
            if parametr == "all":
                if count_vacancies.__contains__(vacancy.published_at):
                    count_vacancies[vacancy.published_at] += 1
                else:
                    count_vacancies[vacancy.published_at] = 1
        if (len(count_vacancies) == 0):
            return {2022: 0}
        return count_vacancies

    def get_count_vacancies_by_profession(data: DataSet, profession):
        """Считает количество вакансий по заданной профессии.

        :param profession: название профессии
        :return:
            dict: словарь, где
                                ключ: год
                                значение: количество вакансий заданной профессии за этот год
        """
        count_vacancies_by_profession = dict()
        for vacancy in data.vacancies_objects:
            if vacancy.name.__contains__(profession):
                if not count_vacancies_by_profession.__contains__(vacancy.published_at):
                    count_vacancies_by_profession[vacancy.published_at] = 1
                else:
                    count_vacancies_by_profession[vacancy.published_at] += 1
        if (len(count_vacancies_by_profession) == 0):
            return {2022: 0}
        return count_vacancies_by_profession

    def get_general_salary_level_by_year(data: DataSet, parametr):
        """Считает зарплату всех вакансий.

        :param parametr: параметр, определяющий необходимость подсчета зарплаты всех вакансий
        :return:
            dict: словарь, где
                                ключ: год
                                значение: зарплата за этот год
        """
        general_salary_level_by_year = dict()
        for vac in data.vacancies_objects:
            if general_salary_level_by_year.__contains__(vac.published_at):
                general_salary_level_by_year[vac.published_at] += InputConnect.get_right_course(vac)
            else:
                general_salary_level_by_year[vac.published_at] = InputConnect.get_right_course(vac)
        if (len(general_salary_level_by_year) == 0):
            return {2022: 0}
        for key in general_salary_level_by_year.keys():
            if parametr == "all":
                general_salary_level_by_year[key] = math.floor(
                    general_salary_level_by_year[key] / data.general_count_vacancies_by_year[key])
        return general_salary_level_by_year

    def get_salary_level_by_profession(data: DataSet, profession):
        """Считает зарплату для заданной профессии.

        :param profession: название профессии
        :return:
            dict: словарь, где
                                ключ: год
                                значение: зарплата заданной профессии за этот год
        """
        salary_level_by_profession = dict()
        for vacancy in data.vacancies_objects:
            if vacancy.name.__contains__(profession):
                if salary_level_by_profession.__contains__(vacancy.published_at):
                    salary_level_by_profession[vacancy.published_at] += InputConnect.get_right_course(vacancy)
                else:
                    salary_level_by_profession[vacancy.published_at] = InputConnect.get_right_course(vacancy)
        if (len(salary_level_by_profession) == 0):
            return {2022: 0}
        for key in salary_level_by_profession.keys():
            if profession != "all":
                salary_level_by_profession[key] = math.floor(
                    salary_level_by_profession[key] / data.count_vacancies_by_profession[key])
        return salary_level_by_profession

    def get_proportion_vacancy_by_cities(data: DataSet):
        """Считает долю вакансий в городах, в которых кол-во вакансий больше или равно 1% от общего числа вакансий

        :param data: объект класса DataSet
        :return:
            dict: словарь, где
                                ключ: город
                                значение: доля вакансий в этом городе
        """
        proportion_vacancy_by_cities = dict()
        for vacancy in data.vacancies_objects:
            if proportion_vacancy_by_cities.__contains__(vacancy.area_name):
                proportion_vacancy_by_cities[vacancy.area_name] += 1
            else:
                proportion_vacancy_by_cities[vacancy.area_name] = 1
        return proportion_vacancy_by_cities

    def get_salary_level_by_cities(data: DataSet):
        """Считает уровень зарплаты в городах, в которых кол-во вакансий больше или равно 1% от общего числа вакансий

        :param data: объект класса DataSet
        :return:
            dict: словарь, где
                                ключ: город
                                значение: уровень зарплаты в этом городе
        """
        salary_level_by_cities = dict()
        for vacancy in data.vacancies_objects:
            if math.floor(
                    data.proportion_vacancy_by_cities[vacancy.area_name] / len(data.vacancies_objects) * 100) >= 1:
                if salary_level_by_cities.__contains__(vacancy.area_name):
                    salary_level_by_cities[vacancy.area_name] += InputConnect.get_right_course(vacancy)
                else:
                    salary_level_by_cities[vacancy.area_name] = InputConnect.get_right_course(vacancy)
        for key in salary_level_by_cities:
            salary_level_by_cities[key] = math.floor(
                salary_level_by_cities[key] / data.proportion_vacancy_by_cities[key])
        result = {k: v for k, v in sorted(salary_level_by_cities.items(), key=lambda item: item[1], reverse=True)}
        return result

    def get_right_proportion_vacancy_by_cities(data: DataSet):
        """Пересчитывает верную долю вакансий в городах, в которых кол-во вакансий больше или равно 1% от общего числа вакансий

        :param data: объект класса DataSet
        :return:
            dict: словарь, где
                                ключ: город
                                значение: доля вакансий в этом городе
        """
        data.proportion_vacancy_by_cities = {k: round(v / len(data.vacancies_objects), 4)
                                             for k, v in data.proportion_vacancy_by_cities.items()}
        data.proportion_vacancy_by_cities = {k: v for k, v in data.proportion_vacancy_by_cities.items()
                                             if math.floor(v * 100 >= 1)}
        return {k: v for k, v in sorted(data.proportion_vacancy_by_cities.items(), key=lambda item: item[1], reverse=True)}

    def print(self, data: DataSet):
        """
        Вычисляет все словари со всей информацией и возращает их для обработки при составлении графика, таблицы и pdf файла.
        :param self: объект класса InputConnect
        :param data: объект класса DataSet

        :return:
            dict: Динамика уровня зарплат по годам
            dict: Динамика количества вакансий по годам
            dict: Динамика уровня зарплат по годам для выбранной профессии
            dict: Динамика количества вакансий по годам для выбранной профессии
            dict: Уровень зарплат по городам (в порядке убывания) - только первые 10 значений
            dict: Доля вакансий по городам (в порядке убывания) - только первые 10 значений
        """
        data.general_count_vacancies_by_year = InputConnect.get_count_vacancies(data, "all")
        data.general_salary_level_by_year = InputConnect.get_general_salary_level_by_year(data, "all")

        data.count_vacancies_by_profession = InputConnect.get_count_vacancies_by_profession(data, self.prof_name_init)
        data.salary_level_by_profession = InputConnect.get_salary_level_by_profession(data, self.prof_name_init)

        data.proportion_vacancy_by_cities = InputConnect.get_proportion_vacancy_by_cities(data)
        data.salary_level_by_cities = InputConnect.get_salary_level_by_cities(data)

        data.salary_level_by_cities_first_ten = dict(list(data.salary_level_by_cities.items())[:10])

        data.proportion_vacancy_by_cities = InputConnect.get_right_proportion_vacancy_by_cities(data)
        data.proportion_vacancy_by_cities_first_ten = dict(list(data.proportion_vacancy_by_cities.items())[:10])

        return data.general_salary_level_by_year, data.general_count_vacancies_by_year, data.salary_level_by_profession, data.count_vacancies_by_profession, data.salary_level_by_cities_first_ten, data.proportion_vacancy_by_cities_first_ten

class Report(InputConnect):
    """Класс для формирования отчетности: Excel-таблицы, графиков и общего отчета в виде pdf-файла.

    Attribures:
        InputConnect: объект класса InputConnect.
    """
    def __init__(self, dict1, dict2, dict3, dict4 ,dict5, dict6):
        """Инициализирует объект Report.

        Args:
            dict1 (dict): словарь с динамикой уровня зарплат по годам
            dict2 (dict): словарь с динамикой количества вакансий по годам
            dict3 (dict): словарь с динамикой уровня зарплат по годам для выбранной профессии
            dict4 (dict): словарь с динамикой количества вакансий по годам для выбранной профессии
            dict5 (dict): словарь с уровенем зарплат по городам (в порядке убывания) - только первые 10 значений
            dict6 (dict): словарь с долей вакансий по городам (в порядке убывания) - только первые 10 значений
        """
        self.general_salary_level_by_year = dict1
        self.general_count_vacancies_by_year = dict2
        self.salary_level_by_profession = dict3
        self.count_vacancies_by_profession = dict4
        self.salary_level_by_cities_first_ten = dict5
        self.proportion_vacancy_by_cities_first_ten = dict6

    def as_text(val):
        """Преобразует значение в строку (str)

        :return:
            str: если значение не None
        """
        if val is None:
            return ""
        return str(val)

    def generate_excel(self):
        """Генерирует отчет в виде excel-файла.
            На первом листе отображается статистика по годам, на втором листе - статистика по городам.
        """
        wb = Workbook()
        sheet1 = wb.active
        thin = Side(border_style='thin', color='000000')
        sheet1.title = 'Статистика по годам'
        sheet2 = wb.create_sheet('Статистика по городам')
        prof = input_profession_name

        border = Border(left=thin, top=thin, right=thin, bottom=thin)

        heads1 = ['Год', 'Средняя зарплата',
                  f'Средняя зарплата - {prof}',
                  'Количество вакансий по годам', f'Количество вакансий по годам - {prof}']

        for i, head in enumerate(heads1):
            sheet1.cell(row = 1, column = (i + 1), value = head).font = Font(bold=True)

        year = list(self.general_salary_level_by_year.keys())[0]
        for i in range(len(self.general_salary_level_by_year)):
            sheet1.append([year, self.general_salary_level_by_year[year], self.salary_level_by_profession[year],
                           self.general_count_vacancies_by_year[year], self.count_vacancies_by_profession[year]])
            year += 1

        for column in sheet1.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        for column in sheet1.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet1.column_dimensions[column[1].column_letter].width = length + 2

        heads2 = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий по городам']

        for i, head in enumerate(heads2):
            sheet2.cell(row = 1, column = (i + 1), value = head).font = Font(bold=True)

        for i in range(len(self.salary_level_by_cities_first_ten)):
            city = list(self.salary_level_by_cities_first_ten.keys())[i]
            sheet2.append([city, self.salary_level_by_cities_first_ten[city]])

        j = 2
        for city in self.proportion_vacancy_by_cities_first_ten.keys():
            sheet2[f'D{j}'].value = city
            sheet2[f'E{j}'].value = f'{round(self.proportion_vacancy_by_cities_first_ten[city] * 100, 2)}%'
            sheet2[f'E{j}'].alignment = openpyxl.styles.Alignment(horizontal='right')
            j += 1

        for column in sheet2.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 2

        for i in range(1, 12):
            sheet2[f'A{i}'].border = border
            sheet2[f'B{i}'].border = border
            sheet2[f'D{i}'].border = border
            sheet2[f'E{i}'].border = border

        wb.save('report.xlsx')

    def generate_image(self):
        """Генерирует отчет в виде изображения с графиками, сохраняя его в .png файл.
        """
        fig = plt.figure()

        ax_1 = fig.add_subplot(2, 2, 1)
        ax_2 = fig.add_subplot(2, 2, 2)
        ax_3 = fig.add_subplot(2, 2, 3)
        ax_4 = fig.add_subplot(2, 2, 4)

        labels = self.general_salary_level_by_year.keys()

        profession = input_profession_name
        x = np.arange(len(labels))

        ax_1.set_title('Уровень зарплат по годам', fontsize=8)
        ax_1.bar(x, self.general_salary_level_by_year.values(), width = 0.4, label="средняя з/п")
        ax_1.bar(x + 0.4, self.salary_level_by_profession.values(), width = 0.4, label=f'з/п {profession.lower()}')
        ax_1.set_xticks(x, labels, rotation = 90)
        ax_1.legend(fontsize=8, loc=2)
        ax_1.tick_params(axis='both', labelsize=8)
        ax_1.grid(axis='y')

        ax_2.set_title('Количество вакансий по годам', fontsize=8)
        ax_2.bar(x, self.general_count_vacancies_by_year.values(), width = 0.4, label = "Количество вакансий")
        ax_2.bar(x + 0.4, self.count_vacancies_by_profession.values(), width = 0.4, label = f'Количество вакансий {profession.lower()}')
        ax_2.set_xticks(x, labels, rotation = 90)
        ax_2.legend(fontsize = 8, loc = 2)
        ax_2.tick_params(axis='both', labelsize = 8)
        ax_2.grid(axis='y')

        cities = [city.replace('-', '-\n').replace(' ', '\n') for city in self.salary_level_by_cities_first_ten.keys()]

        ax_3.set_title('Уровень зарплат по городам', fontsize=8)
        ax_3.barh(cities, list(reversed(self.salary_level_by_cities_first_ten.values())), height=0.8, align = 'center')
        ax_3.xaxis.set_tick_params(labelsize = 8)
        ax_3.yaxis.set_tick_params(labelsize = 6)
        ax_3.grid(axis='x')

        labels_cities =  list(self.proportion_vacancy_by_cities_first_ten.keys()) + ['Другие']
        others_sizes = [1 - sum(self.proportion_vacancy_by_cities_first_ten.values())]

        ax_4.set_title('Доля вакансий по городам', fontsize=8)
        ax_4.pie(list(self.proportion_vacancy_by_cities_first_ten.values()) + others_sizes, labels=labels_cities, textprops={'fontsize': 6})
        ax_4.axis('equal')

        plt.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self):
        """Генерирует отчет с графиками и таблицами в виде .pdf файла."""
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('pdf.html')

        profession = input_profession_name
        image_file = 'C:\\Users\\kh_ju\\homework python\\Kharlamkova\\graph.png'

        rows = []
        for year in general_salary_level_by_year.keys():
            rows.append([year, general_salary_level_by_year[year], salary_level_by_profession[year],
                         general_count_vacancies_by_year[year], count_vacancies_by_profession[year]])

        for key, value in proportion_vacancy_by_cities_first_ten.items():
            value = f'{round(value * 100, 2)}%'
            proportion_vacancy_by_cities_first_ten[key] = value

        pdf_template = template.render({'profession': profession, 'image_file': image_file, 'rows': rows,
                                        'general_salary_level_by_year': general_salary_level_by_year,
                                        'general_count_vacancies_by_year': general_count_vacancies_by_year,
                                        'salary_level_by_profession': salary_level_by_profession,
                                        'count_vacancies_by_profession': count_vacancies_by_profession,
                                        'salary_level_by_cities_first_ten': salary_level_by_cities_first_ten,
                                        'proportion_vacancy_by_cities_first_ten': proportion_vacancy_by_cities_first_ten})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Users\kh_ju\homework python\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

if __name__ == '__main__':
    input_file_name = input("Введите название файла: ")
    input_profession_name = input("Введите название профессии: ")
    user_input = input("Статистика или вакансии? ").lower()
    profile = cProfile.Profile()
    profile.enable()
    dataset_vacancies = DataSet.get_dataset(file_name=input_file_name)
    InputConnect.print(InputConnect(input_file_name=input_file_name, input_profession_name=input_profession_name),
                       dataset_vacancies)

    dicts = InputConnect.print(InputConnect(input_file_name=input_file_name, input_profession_name=input_profession_name),
                               dataset_vacancies)

    general_salary_level_by_year = dicts[0]
    general_count_vacancies_by_year = dicts[1]
    salary_level_by_profession = dicts[2]
    count_vacancies_by_profession = dicts[3]
    salary_level_by_cities_first_ten = dicts[4]
    proportion_vacancy_by_cities_first_ten = dicts[5]

    report = Report(dict1=general_salary_level_by_year, dict2=general_count_vacancies_by_year,
                                 dict3=salary_level_by_profession, dict4=count_vacancies_by_profession,
                                 dict5=salary_level_by_cities_first_ten, dict6=proportion_vacancy_by_cities_first_ten)

    if user_input == 'вакансии':
        Report.generate_image(report)
    elif user_input == 'статистика':
        Report.generate_excel(report)
    profile.disable()
    profile.print_stats(1)




